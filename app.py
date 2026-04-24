import streamlit as st

import json
import csv
import io
import os
import tempfile
import openpyxl
from openpyxl import load_workbook
from groq import Groq

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="StayVista Acquisition Tool",
    page_icon="🏡",
    layout="wide"
)

# ── Header ─────────────────────────────────────────────────────────────────────
st.title("🏡 StayVista Property Acquisition Tool")
st.caption("Upload a property walkthrough video → get a structured summary + CSV instantly")
st.divider()

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Configuration")
    groq_api_key = st.text_input(
        "Groq API Key",
        type="password",
        placeholder="gsk_...",
        help="Free at console.groq.com"
    )
    st.markdown("[Get free Groq API key →](https://console.groq.com)")
    st.divider()
    st.markdown("**Supported formats**")
    st.markdown("MP4, MOV, MKV, AVI, M4A, MP3, WAV")
    st.markdown("**Supported languages**")
    st.markdown("Hindi, English, Hinglish — auto-detected")



# ── System prompt ──────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are a real estate data extraction assistant for StayVista,
a luxury villa rental platform. Given a transcription of a property walkthrough video,
return a JSON object with EXACTLY these two keys:

1. "summary": Write a clean, structured property summary using ONLY the sections below.
   Include a section ONLY if it was mentioned in the transcription — skip it entirely if not.
   Do not add assumptions or generic filler. Every line must come from the transcription.

   Format the summary exactly like this (use only relevant sections):

   PROPERTY OVERVIEW
   [Property name, type, location, gated community or not, age/condition if mentioned]

   ROOMS & LAYOUT
   [Total rooms, bedrooms, bathrooms, floors, special rooms like pooja/study/theatre]

   AMENITIES & FEATURES
   [Pools (count + type), jacuzzi, gym, spa, garden, terrace, BBQ, bonfire area, etc.]

   KITCHEN & INTERIORS
   [Kitchen type, furnishing status, AC, interiors highlights]

   OUTDOOR & VIEWS
   [View type, plot area, garden size, outdoor features]

   SERVICES & SECURITY
   [Parking, caretaker, security staff, CCTV, WiFi, generator, EV charging]

   APPROVALS & LEGAL
   [RERA approval, RERA number, any other approvals]

   PRICING
   [Monthly rent, nightly rate, security deposit, minimum stay]

   POC NOTES
   [Any additional observations from the field agent]

2. "csv_data": A flat dict with ALL these keys (use null only if genuinely not mentioned):
   property_name, location, city, state, is_gated_community, community_name,
   nearby_landmarks, property_type, total_rooms, bedrooms, bathrooms, total_floors,
   area_sqft, plot_area_sqft, furnishing_status, age_of_property_years,
   is_rera_approved, rera_number, other_approvals,
   has_swimming_pool, number_of_swimming_pools, pool_type, has_jacuzzi,
   number_of_jacuzzis, has_private_pool, has_gym, has_spa, has_sauna,
   has_game_room, has_home_theatre, has_kids_play_area, has_garden,
   garden_area_sqft, has_terrace, has_balcony, has_barbeque_area,
   has_bonfire_area, view_type, kitchen_type, has_dining_area, has_living_room,
   has_study_room, has_pooja_room, has_parking, parking_capacity, has_caretaker,
   has_security_staff, has_cctv, has_wifi, has_ac, has_generator_backup,
   has_ev_charging, monthly_rent_inr, nightly_rate_inr, security_deposit_inr,
   minimum_stay_nights, special_features, condition_of_property, acquisition_poc_notes

Return ONLY valid JSON. No markdown. Never invent details."""

# ── Excel helper ───────────────────────────────────────────────────────────────
EXCEL_FILE = "stayvista_acquisitions.xlsx"

def save_to_excel(csv_data):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        existing_headers = [cell.value for cell in ws[1]]
        # Add any new columns not previously seen
        for key in csv_data.keys():
            display = key.replace("_", " ").title()
            if display not in existing_headers:
                ws.cell(row=1, column=len(existing_headers) + 1, value=display)
                existing_headers.append(display)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Property Acquisitions"
        existing_headers = [k.replace("_", " ").title() for k in csv_data.keys()]
        for col_num, header in enumerate(existing_headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="6C3FC5")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", wrap_text=True)

    # Map headers back to keys and append row
    header_to_key = {k.replace("_", " ").title(): k for k in csv_data.keys()}
    new_row = [csv_data.get(header_to_key.get(h)) for h in existing_headers]
    ws.append(new_row)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)
    return ws.max_row - 1  # total properties saved

# ── Main app ───────────────────────────────────────────────────────────────────
if not groq_api_key:
    st.info("👈 Enter your Groq API key in the sidebar to get started.")
    st.stop()

uploaded_file = st.file_uploader(
    "Upload property walkthrough video or audio",
    type=["mp4", "mov", "mkv", "avi", "m4a", "mp3", "wav"]
)

if uploaded_file:
    if uploaded_file.name.endswith(("mp4", "mov", "mkv", "avi")):
        st.video(uploaded_file)
    else:
        st.audio(uploaded_file)

    if st.button("🚀 Process Property", type="primary", use_container_width=True):

        suffix = os.path.splitext(uploaded_file.name)[-1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(uploaded_file.read())
            tmp_path = tmp.name

        try:
# ── Step 1: Transcribe via Groq Whisper API ───────────────────────────────
          with st.status("🎙️ Transcribing audio...", expanded=True) as status:
             st.write("Sending to Groq Whisper API...")
             client = Groq(api_key=groq_api_key)

           with open(tmp_path, "rb") as audio_file:
            transcription_response = client.audio.transcriptions.create(
            file=(uploaded_file.name, audio_file.read()),
            model="whisper-large-v3",
            response_format="text",
            language="hi",        # set to "hi" for Hindi, or remove for auto-detect
            )

    transcription = transcription_response
    status.update(label="✅ Transcription done", state="complete")
    # ── Step 1b: Translate to English if needed ───────────────────────────────
           with st.status("🌐 Translating to English...", expanded=True) as status:
              st.write("Detecting language and translating...")
            translation_response = client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[
                        {"role": "system", "content": 
                        "You are a translator. If the text is in Hindi or Hinglish, "
                        "translate it to English. If it is already in English, return it as-is. "
                        "Return ONLY the translated text, nothing else."},
                        {"role": "user", "content": transcription}
                   ],
                   temperature=0.1
                   )
            transcription = translation_response.choices[0].message.content
            status.update(label="✅ Translation done", state="complete")

            with st.expander("📄 View transcription"):
              st.write(transcription)

            # ── Step 2: AI Analysis ────────────────────────────────────────────
            with st.status("🤖 Analysing with Llama 3...", expanded=True) as status:
                st.write("Extracting property details...")
                client = Groq(api_key=groq_api_key)
                response = client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": f"Transcription:\n\n{transcription}"}
                    ],
                    temperature=0.1,
                    response_format={"type": "json_object"}
                )
                raw = response.choices[0].message.content
                try:
                    data = json.loads(raw)
                except json.JSONDecodeError:
                    clean = raw.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
                    data = json.loads(clean)
                status.update(label="✅ Analysis complete", state="complete")

            summary  = data["summary"]
            csv_data = data["csv_data"]

            # ── Step 3: Display results ────────────────────────────────────────
            st.divider()
            col1, col2 = st.columns(2, gap="large")

            with col1:
                st.subheader("📋 Property Summary")
                for line in summary.strip().split("\n"):
                    if line.strip().isupper() and len(line.strip()) > 2:
                        st.markdown(f"**{line.strip()}**")
                    elif line.strip():
                        st.markdown(line)

            with col2:
                st.subheader("📊 Extracted Data")
                found = {k: v for k, v in csv_data.items() if v is not None}
                st.caption(f"{len(found)} of {len(csv_data)} fields captured")

                groups = {
                    "📍 Location":  ["property_name","location","city","state",
                                     "is_gated_community","community_name","nearby_landmarks"],
                    "🏠 Property":  ["property_type","total_rooms","bedrooms","bathrooms",
                                     "total_floors","area_sqft","plot_area_sqft",
                                     "furnishing_status","age_of_property_years","condition_of_property"],
                    "✅ Approvals": ["is_rera_approved","rera_number","other_approvals"],
                    "🏊 Amenities": ["has_swimming_pool","number_of_swimming_pools","pool_type",
                                     "has_jacuzzi","number_of_jacuzzis","has_private_pool",
                                     "has_gym","has_spa","has_sauna","has_game_room",
                                     "has_home_theatre","has_kids_play_area","has_garden",
                                     "garden_area_sqft","has_terrace","has_balcony",
                                     "has_barbeque_area","has_bonfire_area","view_type"],
                    "🛋️ Interiors": ["kitchen_type","has_dining_area","has_living_room",
                                     "has_study_room","has_pooja_room"],
                    "🔒 Services":  ["has_parking","parking_capacity","has_caretaker",
                                     "has_security_staff","has_cctv","has_wifi","has_ac",
                                     "has_generator_backup","has_ev_charging"],
                    "💰 Pricing":   ["monthly_rent_inr","nightly_rate_inr",
                                     "security_deposit_inr","minimum_stay_nights"],
                    "📝 Notes":     ["special_features","acquisition_poc_notes"],
                }

                for group_name, keys in groups.items():
                    group_data = {k: csv_data[k] for k in keys
                                  if k in csv_data and csv_data[k] is not None}
                    if group_data:
                        with st.expander(group_name, expanded=True):
                            for k, v in group_data.items():
                                label = k.replace("_", " ").title()
                                st.markdown(f"**{label}:** {v}")

            # ── Step 4: Save to Excel + Downloads ─────────────────────────────
            st.divider()
            st.subheader("⬇️ Download Outputs")

            total_saved = save_to_excel(csv_data)
            st.success(f"✅ Property saved! Total properties in Excel: **{total_saved}**")

            dl1, dl2 = st.columns(2)

            prop_name = (csv_data.get("property_name") or "property").replace(" ", "_")

            with dl1:
                st.download_button(
                    label="📄 Download Summary (.txt)",
                    data=summary,
                    file_name=f"{prop_name}_summary.txt",
                    mime="text/plain",
                    use_container_width=True
                )

            with dl2:
                with open(EXCEL_FILE, "rb") as f:
                    st.download_button(
                        label=f"📊 Download Excel ({total_saved} properties)",
                        data=f,
                        file_name=EXCEL_FILE,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

        finally:
            os.unlink(tmp_path)
